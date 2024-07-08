import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service as ChromeService
import sys
import os
import time

# Ensure the parent directory is in the system path for module imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from login_page import OrangeLoginPage


@pytest.fixture
def browser():
    chrome_service = ChromeService(r"C:\Users\HP\OneDrive\Desktop\chromedriver.exe")
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()
    yield driver
    # time.sleep(6)
    driver.quit()


def read_test_data_from_excel(file_path=r"C:\Users\HP\OneDrive\Documents\Book1.xlsx",
                              sheet_name='Sheet1'):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Ensure both username and password are present
            test_data.append((row[0], row[1]))
        else:
            print(f"Skipping invalid row: {row}")

    return test_data


def test_orange_login(browser, username, password):
    login_page = OrangeLoginPage(browser)

    try:
        # Navigate to the Amazon login page
        login_page.navigate_to_login_page()

        # Wait for the sign-in button to be visible
        # WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.sign_in_button))

        # Perform login
        # login_page.click_sign_in_button()

        # Wait for username input field to be visible
        WebDriverWait(browser, 30).until(EC.element_to_be_clickable(login_page.username_input))

        login_page.enter_username(username)
        login_page.continue_username()
        # If "Continue" button is present, click it

        # Wait for password input field to be visible
        WebDriverWait(browser, 30).until(EC.element_to_be_clickable(login_page.password_input))

        login_page.enter_password(password)
        login_page.continue_password()
        login_page.click_login_button()

        # Wait for login success (customize as per your application)
        WebDriverWait(browser, 30).until(EC.title_contains("Your orange"))
    except Exception as e:
        print(f"Error during login process: {e}")
        raise


@pytest.mark.parametrize("username,password", read_test_data_from_excel())
def test_orange_login_data_driven(browser, username, password):
    read_test_data_from_excel()
    test_orange_login(browser, username, password)


if __name__ == "__main__":
    pytest.main(["-v", "test_orange_login.py"])