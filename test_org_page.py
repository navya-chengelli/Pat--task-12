

from org_login_page import OrangeHRMLoginPage
import pytest
import openpyxl
# from test_data.xlsx import test_data

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import os


file_path = os.path.abspath("testdata.xlsx")

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 7:
            row = tuple(list(row) + [None] * (7 - len(row)))  # Pad the row with None to ensure it has 7 elements
        data.append((row[0], row[1], row[2], row[3], row[4], row[5], row[6]))
    return data, wb, ws

def write_result(ws, row, result):
    ws.cell(row=row, column=7, value=result)
    now = datetime.now()
    ws.cell(row=row, column=4, value=now.date())
    ws.cell(row=row, column=5, value=now.time())

@pytest.fixture
def browser():
    # Update this path to the actual location of your ChromeDriver executable
    chrome_service = ChromeService(r"C:\path\to\your\chromedriver.exe")
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()
    yield driver
    driver.quit()

# Read data from Excel and ensure it includes all 7 columns
test_data, workbook, worksheet = read_excel(file_path)

@pytest.mark.parametrize("test_id, username, password, date, time, tester, result", test_data)
def test_orangehrm_login(browser, test_id, username, password, date, time, tester, result):
    login_page = OrangeHRMLoginPage(browser)
    login_page.navigate_to_login_page()

    try:
        login_page.login(username, password)

        # Wait for login success (customize as per your application)
        WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.ID, "dashboard-quick-launch-panel-menu_holder"))
        )
        test_result = "Passed"
    except Exception as e:
        print(f"Error during login: {e}")
        test_result = "Failed"

    # Write the result back to the Excel file
    write_result(worksheet, test_id + 1, test_result)
    workbook.save(file_path)

if __name__ == "__main__":
    pytest.main(["-v", __file__])




